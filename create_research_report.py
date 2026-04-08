"""
Professional Research Report Generator
AI Stock Prediction System — V3 Pipeline Analysis
100 NSE Stocks, Walk-Forward Validation
"""

import os, sys, json
import pandas as pd
import numpy as np
from pathlib import Path
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from datetime import datetime

# ── Paths ──────────────────────────────────────────────────────────────
RUN_DIR = Path("V3/06_results/runs/20260307_141956")
OUT_FILE = RUN_DIR / "RESEARCH_ANALYSIS_100_STOCKS.xlsx"

# ── Load Data ──────────────────────────────────────────────────────────
print("Loading data...")
detail_df = pd.read_csv(RUN_DIR / "all_windows_detail.csv")
summary_df = pd.read_csv(RUN_DIR / "summary.csv")
preds_df   = pd.read_csv(RUN_DIR / "next_day_predictions.csv")

MODEL_COLS = [
    "lgbm_acc", "xgb_acc", "lstm_acc", "bilstm_acc", "gru_acc",
    "cnn_lstm_acc", "cnn_gru_acc", "tcn_gru_acc", "tcn_transformer_acc", "nbeats_acc"
]
MODEL_NAMES = [
    "LightGBM", "XGBoost", "LSTM", "BiLSTM", "GRU",
    "CNN-LSTM", "CNN-GRU", "TCN-GRU", "TCN-Transformer", "N-BEATS"
]
AVG_MODEL_COLS = [f"avg_{c}" for c in MODEL_COLS]

# ── Style Helpers ──────────────────────────────────────────────────────
def hex_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def mk_border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def mk_font(bold=False, size=11, color="000000", italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic, name="Calibri")

def mk_align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

HEADER_FILL   = hex_fill("1F3864")   # dark navy
SUBHEAD_FILL  = hex_fill("2E75B6")   # medium blue
ALT_FILL      = hex_fill("EBF3FB")   # light blue
WHITE_FILL    = hex_fill("FFFFFF")
GREEN_FILL    = hex_fill("C6EFCE")
YELLOW_FILL   = hex_fill("FFEB9C")
RED_FILL      = hex_fill("FFC7CE")
ORANGE_FILL   = hex_fill("FCE4D6")
GOLD_FILL     = hex_fill("FFD966")

HEADER_FONT   = mk_font(bold=True, size=11, color="FFFFFF")
TITLE_FONT    = mk_font(bold=True, size=14, color="1F3864")
SUBTITLE_FONT = mk_font(bold=False, size=11, color="44546A", italic=True)
BODY_FONT     = mk_font(size=10)
BOLD_FONT     = mk_font(bold=True, size=10)
THIN_BORDER   = mk_border("thin")

def style_header_row(ws, row, ncols, fill=HEADER_FILL, font=HEADER_FONT):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = mk_align("center", "center", wrap=True)
        cell.border = THIN_BORDER

def style_data_row(ws, row, ncols, fill=None, font=BODY_FONT):
    f = fill or (ALT_FILL if row % 2 == 0 else WHITE_FILL)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = f
        cell.font = font
        cell.alignment = mk_align("center", "center")
        cell.border = THIN_BORDER

def set_col_width(ws, col_widths):
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

def freeze_and_filter(ws, freeze="A2", autofilter=None):
    ws.freeze_panes = freeze
    if autofilter:
        ws.auto_filter.ref = autofilter

def pct(val, decimals=2):
    if pd.isna(val):
        return "N/A"
    return f"{val*100:.{decimals}f}%"

def color_for_acc(acc):
    if pd.isna(acc):
        return WHITE_FILL
    if acc >= 0.60:
        return GREEN_FILL
    elif acc >= 0.55:
        return hex_fill("C6EFCE")
    elif acc >= 0.50:
        return YELLOW_FILL
    elif acc >= 0.45:
        return ORANGE_FILL
    else:
        return RED_FILL

# ════════════════════════════════════════════════════════════════════════
# Precompute analytics
# ════════════════════════════════════════════════════════════════════════
print("Computing analytics...")

# Per-symbol best model & window
sym_best = {}
for sym, grp in detail_df.groupby("symbol"):
    best_row = None
    best_acc = -1
    for _, row in grp.iterrows():
        for mc in MODEL_COLS:
            if row[mc] > best_acc:
                best_acc = row[mc]
                best_row = (row["window_id"], mc, best_acc)
    sym_best[sym] = best_row  # (window_id, model_col, acc)

# Per-model averages across all windows all stocks
model_avgs = {}
for mc, mn in zip(MODEL_COLS, MODEL_NAMES):
    vals = detail_df[mc].dropna()
    model_avgs[mn] = {
        "mean": vals.mean(),
        "std":  vals.std(),
        "min":  vals.min(),
        "max":  vals.max(),
        "pct60": (vals >= 0.60).mean(),
        "pct55": (vals >= 0.55).mean(),
        "pct50": (vals >= 0.50).mean(),
    }

# Per-window averages
window_avgs = {}
for wid, grp in detail_df.groupby("window_id"):
    window_avgs[wid] = {
        "oos_acc": grp["oos_accuracy"].mean(),
        "model_avgs": {mn: grp[mc].mean() for mc, mn in zip(MODEL_COLS, MODEL_NAMES)},
        "train_ratio": grp["train_ratio"].iloc[0],
    }

# Overall stats
overall_oos      = summary_df["oos_accuracy"].mean()
overall_best_mw  = np.mean([v[2] for v in sym_best.values()])
n_above_60       = (summary_df["oos_accuracy"] >= 0.60).sum()
n_above_55       = (summary_df["oos_accuracy"] >= 0.55).sum()
n_above_50       = (summary_df["oos_accuracy"] >= 0.50).sum()
n_below_50       = (summary_df["oos_accuracy"] < 0.50).sum()
n_hit_60_bestmw  = sum(1 for v in sym_best.values() if v[2] >= 0.60)

best_overall_model = max(model_avgs, key=lambda k: model_avgs[k]["mean"])
best_overall_window = max(window_avgs, key=lambda k: window_avgs[k]["oos_acc"])

# ════════════════════════════════════════════════════════════════════════
# Create Workbook
# ════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

# ════════════════════════════════════════════════════════════════════════
# SHEET 1: Executive Summary
# ════════════════════════════════════════════════════════════════════════
print("Sheet 1: Executive Summary...")
ws1 = wb.create_sheet("1. Executive Summary")
ws1.sheet_view.showGridLines = False

# Title block
ws1.merge_cells("A1:H1")
ws1["A1"] = "AI STOCK PREDICTION SYSTEM — V3 PIPELINE RESEARCH ANALYSIS"
ws1["A1"].font = mk_font(bold=True, size=16, color="1F3864")
ws1["A1"].alignment = mk_align("center", "center")
ws1["A1"].fill = hex_fill("D6E4F0")

ws1.merge_cells("A2:H2")
ws1["A2"] = f"100 NSE Stocks | Walk-Forward Validation | 6 Windows | 10 Models | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
ws1["A2"].font = SUBTITLE_FONT
ws1["A2"].alignment = mk_align("center", "center")

ws1.row_dimensions[1].height = 30
ws1.row_dimensions[2].height = 20

# Key metrics section
ws1.merge_cells("A4:H4")
ws1["A4"] = "KEY PERFORMANCE METRICS"
ws1["A4"].font = mk_font(bold=True, size=12, color="FFFFFF")
ws1["A4"].fill = HEADER_FILL
ws1["A4"].alignment = mk_align("center", "center")

metrics = [
    ("Stocks Analyzed",           f"100",                                   "Total NSE stocks in Nifty 100 universe"),
    ("Average OOS Accuracy",      f"{overall_oos*100:.2f}%",                "Ensemble accuracy across all stocks & windows"),
    ("Best Model/Window Avg",     f"{overall_best_mw*100:.2f}%",            "Per-stock best model × best window average"),
    ("60% Target Met (Best M/W)", f"{n_hit_60_bestmw}/100 stocks",          "Stocks achieving ≥60% with optimal model/window"),
    ("Stocks ≥60% (Ensemble)",    f"{n_above_60}/100 stocks",               "Stocks where ensemble OOS ≥ 60%"),
    ("Stocks ≥55% (Ensemble)",    f"{n_above_55}/100 stocks",               "Stocks where ensemble OOS ≥ 55%"),
    ("Stocks ≥50% (Ensemble)",    f"{n_above_50}/100 stocks",               "Stocks with above-random ensemble accuracy"),
    ("Stocks <50% (Ensemble)",    f"{n_below_50}/100 stocks",               "Stocks below random baseline — need attention"),
    ("Best Model Overall",        best_overall_model,                        f"Avg: {model_avgs[best_overall_model]['mean']*100:.2f}%"),
    ("Best Window Overall",       f"Window {best_overall_window}",          f"Train ratio: {window_avgs[best_overall_window]['train_ratio']*100:.0f}% | Avg OOS: {window_avgs[best_overall_window]['oos_acc']*100:.2f}%"),
    ("Active Buy Signals",        f"{(preds_df['action']=='BUY').sum()} stocks", "Next-day BUY signals from production models"),
    ("Active Sell Signals",       f"{(preds_df['action']=='SELL').sum()} stocks","Next-day SELL signals from production models"),
]

fills_m = [GREEN_FILL, hex_fill("C6EFCE"), GREEN_FILL, GREEN_FILL,
           YELLOW_FILL, YELLOW_FILL, hex_fill("E2EFDA"), RED_FILL,
           hex_fill("BDD7EE"), hex_fill("BDD7EE"), hex_fill("E2EFDA"), ORANGE_FILL]

headers_m = ["Metric", "Value", "Notes"]
for c, h in enumerate(headers_m, 1):
    ws1.cell(row=5, column=c).value = h
style_header_row(ws1, 5, 3, SUBHEAD_FILL)

for i, (metric, value, note) in enumerate(metrics, 6):
    ws1.cell(row=i, column=1).value = metric
    ws1.cell(row=i, column=2).value = value
    ws1.cell(row=i, column=3).value = note
    f = fills_m[i-6]
    for c in range(1, 4):
        cell = ws1.cell(row=i, column=c)
        cell.fill = f
        cell.font = mk_font(bold=(c==2), size=10)
        cell.alignment = mk_align("left" if c in (1,3) else "center", "center")
        cell.border = THIN_BORDER
    ws1.row_dimensions[i].height = 18

# Accuracy distribution table
row_off = len(metrics) + 8
ws1.merge_cells(f"A{row_off}:H{row_off}")
ws1[f"A{row_off}"] = "ACCURACY DISTRIBUTION (Ensemble OOS)"
ws1[f"A{row_off}"].font = mk_font(bold=True, size=12, color="FFFFFF")
ws1[f"A{row_off}"].fill = HEADER_FILL
ws1[f"A{row_off}"].alignment = mk_align("center", "center")

dist_headers = ["Range", "Count", "% of Universe", "Classification", "Action"]
for c, h in enumerate(dist_headers, 1):
    ws1.cell(row=row_off+1, column=c).value = h
style_header_row(ws1, row_off+1, 5, SUBHEAD_FILL)

accs = summary_df["oos_accuracy"]
dist_data = [
    ("≥ 65%",     (accs >= 0.65).sum(), GREEN_FILL,  "Excellent",  "Monitor & scale"),
    ("60–64.9%",  ((accs>=0.60)&(accs<0.65)).sum(), GREEN_FILL, "Very Good", "Priority trading"),
    ("55–59.9%",  ((accs>=0.55)&(accs<0.60)).sum(), YELLOW_FILL, "Good",     "Trade with filters"),
    ("50–54.9%",  ((accs>=0.50)&(accs<0.55)).sum(), YELLOW_FILL, "Marginal", "Trade cautiously"),
    ("45–49.9%",  ((accs>=0.45)&(accs<0.50)).sum(), ORANGE_FILL, "Below Random", "Do not trade"),
    ("< 45%",     (accs < 0.45).sum(),               RED_FILL,   "Poor",     "Exclude / retrain"),
]
for i, (rng, cnt, fill, cls, act) in enumerate(dist_data, row_off+2):
    vals = [rng, cnt, f"{cnt/100*100:.0f}%", cls, act]
    for c, v in enumerate(vals, 1):
        cell = ws1.cell(row=i, column=c)
        cell.value = v
        cell.fill = fill
        cell.font = BODY_FONT
        cell.alignment = mk_align("center", "center")
        cell.border = THIN_BORDER
    ws1.row_dimensions[i].height = 18

set_col_width(ws1, {"A": 28, "B": 20, "C": 18, "D": 20, "E": 25, "F": 18, "G": 18, "H": 18})

# ════════════════════════════════════════════════════════════════════════
# SHEET 2: Per-Stock Full Analysis
# ════════════════════════════════════════════════════════════════════════
print("Sheet 2: Per-Stock Analysis...")
ws2 = wb.create_sheet("2. Per-Stock Analysis")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:V1")
ws2["A1"] = "PER-STOCK FULL ANALYSIS — ALL 100 NSE STOCKS"
ws2["A1"].font = mk_font(bold=True, size=13, color="FFFFFF")
ws2["A1"].fill = HEADER_FILL
ws2["A1"].alignment = mk_align("center", "center")
ws2.row_dimensions[1].height = 25

headers2 = [
    "#", "Symbol", "Status",
    "Ensemble\nOOS Acc", "OOS F1", "Windows", "Predictions",
    "Features", "Rows",
    "LightGBM", "XGBoost", "LSTM", "BiLSTM", "GRU",
    "CNN-LSTM", "CNN-GRU", "TCN-GRU", "TCN-Trans", "N-BEATS",
    "Best Model", "Best Acc",
    "Dir Acc UP", "Dir Acc DN"
]
for c, h in enumerate(headers2, 1):
    ws2.cell(row=2, column=c).value = h
style_header_row(ws2, 2, len(headers2))
ws2.row_dimensions[2].height = 35

summary_df_sorted = summary_df.sort_values("oos_accuracy", ascending=False).reset_index(drop=True)

for i, row in summary_df_sorted.iterrows():
    r = i + 3
    best_info = sym_best.get(row["symbol"], (None, None, None))
    best_mn = dict(zip(MODEL_COLS, MODEL_NAMES)).get(best_info[1], best_info[1]) if best_info[1] else ""

    vals = [
        i+1, row["symbol"], row.get("status","ok"),
        row["oos_accuracy"], row["oos_f1"],
        int(row["n_windows"]), int(row["n_predictions"]),
        int(row.get("n_features",50)), int(row.get("n_rows",0)),
        row["avg_lgbm_acc"], row["avg_xgb_acc"],
        row["avg_lstm_acc"], row["avg_bilstm_acc"], row["avg_gru_acc"],
        row["avg_cnn_lstm_acc"], row["avg_cnn_gru_acc"],
        row["avg_tcn_gru_acc"], row["avg_tcn_transformer_acc"], row["avg_nbeats_acc"],
        best_mn, best_info[2] if best_info[2] else 0,
        row.get("avg_dir_acc_up", 0), row.get("avg_dir_acc_down", 0),
    ]

    for c, v in enumerate(vals, 1):
        cell = ws2.cell(row=r, column=c)
        if isinstance(v, float):
            cell.value = round(v, 4)
            if c in range(4, 22):  # accuracy columns
                cell.number_format = "0.00%"
        else:
            cell.value = v

        # Row fill based on ensemble OOS
        acc_val = row["oos_accuracy"]
        if acc_val >= 0.60:
            base_fill = GREEN_FILL
        elif acc_val >= 0.55:
            base_fill = hex_fill("E2EFDA")
        elif acc_val >= 0.50:
            base_fill = YELLOW_FILL
        elif acc_val >= 0.45:
            base_fill = ORANGE_FILL
        else:
            base_fill = RED_FILL

        cell.fill = base_fill
        cell.font = mk_font(bold=(c==2), size=9)
        cell.alignment = mk_align("center", "center")
        cell.border = THIN_BORDER
    ws2.row_dimensions[r].height = 16

# Column widths
widths2 = {"A":5,"B":13,"C":8,"D":11,"E":9,"F":9,"G":12,"H":10,"I":9,
           "J":11,"K":11,"L":11,"M":11,"N":11,"O":11,"P":11,"Q":11,"R":11,"S":11,
           "T":13,"U":11,"V":11,"W":11}
set_col_width(ws2, widths2)
freeze_and_filter(ws2, "D3", f"A2:W{len(summary_df)+2}")

# Conditional formatting for accuracy columns (D through U)
for col_letter in ["D","J","K","L","M","N","O","P","Q","R","S","U"]:
    ws2.conditional_formatting.add(
        f"{col_letter}3:{col_letter}{len(summary_df)+2}",
        ColorScaleRule(
            start_type="num", start_value=0.40, start_color="FFC7CE",
            mid_type="num",   mid_value=0.50,   mid_color="FFEB9C",
            end_type="num",   end_value=0.65,   end_color="C6EFCE"
        )
    )

# ════════════════════════════════════════════════════════════════════════
# SHEET 3: Model Comparison
# ════════════════════════════════════════════════════════════════════════
print("Sheet 3: Model Comparison...")
ws3 = wb.create_sheet("3. Model Comparison")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:J1")
ws3["A1"] = "MODEL COMPARISON — 10 MODELS ACROSS ALL STOCKS & WINDOWS"
ws3["A1"].font = mk_font(bold=True, size=13, color="FFFFFF")
ws3["A1"].fill = HEADER_FILL
ws3["A1"].alignment = mk_align("center", "center")
ws3.row_dimensions[1].height = 25

headers3 = ["Rank","Model","Avg Accuracy","Std Dev","Min","Max",
            "% ≥60%","% ≥55%","% ≥50%","Classification"]
for c, h in enumerate(headers3, 1):
    ws3.cell(row=2, column=c).value = h
style_header_row(ws3, 2, len(headers3))

sorted_models = sorted(model_avgs.items(), key=lambda x: x[1]["mean"], reverse=True)
for i, (mn, stats) in enumerate(sorted_models, 1):
    r = i + 2
    if stats["mean"] >= 0.55:
        cls = "Strong"
        f = GREEN_FILL
    elif stats["mean"] >= 0.52:
        cls = "Good"
        f = hex_fill("E2EFDA")
    elif stats["mean"] >= 0.50:
        cls = "Marginal"
        f = YELLOW_FILL
    else:
        cls = "Weak"
        f = RED_FILL

    row_vals = [i, mn,
                f"{stats['mean']*100:.3f}%",
                f"{stats['std']*100:.3f}%",
                f"{stats['min']*100:.2f}%",
                f"{stats['max']*100:.2f}%",
                f"{stats['pct60']*100:.1f}%",
                f"{stats['pct55']*100:.1f}%",
                f"{stats['pct50']*100:.1f}%",
                cls]
    for c, v in enumerate(row_vals, 1):
        cell = ws3.cell(row=r, column=c)
        cell.value = v
        cell.fill = f if c >= 3 else (ALT_FILL if i%2==0 else WHITE_FILL)
        cell.font = mk_font(bold=(c in (1,2)), size=10)
        cell.alignment = mk_align("center", "center")
        cell.border = THIN_BORDER
    ws3.row_dimensions[r].height = 20

# Per-stock per-model breakdown
row_off3 = len(sorted_models) + 5
ws3.merge_cells(f"A{row_off3}:K{row_off3}")
ws3[f"A{row_off3}"] = "PER-STOCK MODEL PERFORMANCE MATRIX (Avg across 6 Windows)"
ws3[f"A{row_off3}"].font = mk_font(bold=True, size=12, color="FFFFFF")
ws3[f"A{row_off3}"].fill = SUBHEAD_FILL
ws3[f"A{row_off3}"].alignment = mk_align("center", "center")
ws3.row_dimensions[row_off3].height = 22

m_headers = ["#", "Symbol"] + MODEL_NAMES
for c, h in enumerate(m_headers, 1):
    ws3.cell(row=row_off3+1, column=c).value = h
style_header_row(ws3, row_off3+1, len(m_headers))

for i, row in summary_df_sorted.iterrows():
    r = row_off3 + 2 + i
    vals = [i+1, row["symbol"]] + [row.get(c, 0) for c in AVG_MODEL_COLS]
    for c, v in enumerate(vals, 1):
        cell = ws3.cell(row=r, column=c)
        if isinstance(v, float):
            cell.value = round(v, 4)
            cell.number_format = "0.00%"
            cell.fill = color_for_acc(v)
        else:
            cell.value = v
            cell.fill = ALT_FILL if i%2==0 else WHITE_FILL
        cell.font = mk_font(size=9)
        cell.alignment = mk_align("center", "center")
        cell.border = THIN_BORDER
    ws3.row_dimensions[r].height = 15

set_col_width(ws3, {"A":6,"B":13,"C":13,"D":13,"E":12,"F":12,"G":12,
                    "H":12,"I":12,"J":12,"K":14,"L":14})
freeze_and_filter(ws3, "C2")

# ════════════════════════════════════════════════════════════════════════
# SHEET 4: Window Analysis
# ════════════════════════════════════════════════════════════════════════
print("Sheet 4: Window Analysis...")
ws4 = wb.create_sheet("4. Window Analysis")
ws4.sheet_view.showGridLines = False

ws4.merge_cells("A1:M1")
ws4["A1"] = "WALK-FORWARD WINDOW ANALYSIS — PERFORMANCE BY TRAINING RATIO"
ws4["A1"].font = mk_font(bold=True, size=13, color="FFFFFF")
ws4["A1"].fill = HEADER_FILL
ws4["A1"].alignment = mk_align("center", "center")
ws4.row_dimensions[1].height = 25

win_headers = ["Window","Train Ratio","Avg OOS Acc"] + MODEL_NAMES
for c, h in enumerate(win_headers, 1):
    ws4.cell(row=2, column=c).value = h
style_header_row(ws4, 2, len(win_headers))
ws4.row_dimensions[2].height = 30

for wid in sorted(window_avgs.keys()):
    r = wid + 2
    wa = window_avgs[wid]
    vals = [f"Window {wid}",
            f"{wa['train_ratio']*100:.0f}%",
            f"{wa['oos_acc']*100:.3f}%"] + \
           [f"{wa['model_avgs'][mn]*100:.3f}%" for mn in MODEL_NAMES]
    for c, v in enumerate(vals, 1):
        cell = ws4.cell(row=r, column=c)
        cell.value = v
        if c == 3:
            cell.fill = color_for_acc(wa["oos_acc"])
        elif c > 3:
            mn = MODEL_NAMES[c-4]
            cell.fill = color_for_acc(wa["model_avgs"][mn])
        else:
            cell.fill = ALT_FILL if wid%2==0 else WHITE_FILL
        cell.font = mk_font(bold=(c<=2), size=10)
        cell.alignment = mk_align("center", "center")
        cell.border = THIN_BORDER
    ws4.row_dimensions[r].height = 22

# Per-symbol per-window breakdown
row_off4 = 11
ws4.merge_cells(f"A{row_off4}:N{row_off4}")
ws4[f"A{row_off4}"] = "PER-STOCK WINDOW DETAIL (Ensemble OOS per Window)"
ws4[f"A{row_off4}"].font = mk_font(bold=True, size=12, color="FFFFFF")
ws4[f"A{row_off4}"].fill = SUBHEAD_FILL
ws4[f"A{row_off4}"].alignment = mk_align("center", "center")
ws4.row_dimensions[row_off4].height = 22

w_headers = ["#","Symbol","Win 1\n(70%)","Win 2\n(75%)","Win 3\n(80%)",
             "Win 4\n(85%)","Win 5\n(90%)","Win 6\n(95%)",
             "Best Win","Best Win Acc","Avg","Std Dev","Range","Trend"]
for c, h in enumerate(w_headers, 1):
    ws4.cell(row=row_off4+1, column=c).value = h
style_header_row(ws4, row_off4+1, len(w_headers))
ws4.row_dimensions[row_off4+1].height = 35

for i, sym in enumerate(summary_df_sorted["symbol"], 1):
    r = row_off4 + 1 + i
    sym_data = detail_df[detail_df["symbol"]==sym].sort_values("window_id")
    win_accs = sym_data["oos_accuracy"].tolist()
    # pad if missing windows
    while len(win_accs) < 6:
        win_accs.append(None)

    if all(a is None for a in win_accs):
        best_win, best_win_acc = "N/A", None
    else:
        valid = [(j+1, a) for j,a in enumerate(win_accs) if a is not None]
        best_win, best_win_acc = max(valid, key=lambda x: x[1])

    valid_accs = [a for a in win_accs if a is not None]
    avg_w = np.mean(valid_accs) if valid_accs else None
    std_w = np.std(valid_accs) if len(valid_accs) > 1 else None
    rng_w = (max(valid_accs) - min(valid_accs)) if len(valid_accs) > 1 else None

    # Trend: is window 6 > window 1?
    if win_accs[0] and win_accs[5]:
        trend = "UP" if win_accs[5] > win_accs[0] else ("FLAT" if win_accs[5] == win_accs[0] else "DN")
    else:
        trend = "N/A"

    row_vals = [i, sym] + win_accs + [
        f"Win {best_win}", best_win_acc, avg_w, std_w, rng_w, trend
    ]
    for c, v in enumerate(row_vals, 1):
        cell = ws4.cell(row=r, column=c)
        if v is None:
            cell.value = "N/A"
            cell.fill = hex_fill("D9D9D9")
        elif isinstance(v, float):
            cell.value = round(v, 4)
            cell.number_format = "0.00%"
            if c in range(3, 12):
                cell.fill = color_for_acc(v)
            else:
                cell.fill = ALT_FILL if i%2==0 else WHITE_FILL
        else:
            cell.value = v
            if c == 14:  # trend
                cell.fill = GREEN_FILL if v == "UP" else (RED_FILL if v == "DN" else YELLOW_FILL)
            else:
                cell.fill = ALT_FILL if i%2==0 else WHITE_FILL
        cell.font = mk_font(size=9)
        cell.alignment = mk_align("center", "center")
        cell.border = THIN_BORDER
    ws4.row_dimensions[r].height = 15

set_col_width(ws4, {"A":5,"B":13,"C":11,"D":11,"E":11,"F":11,"G":11,"H":11,
                    "I":11,"J":12,"K":11,"L":10,"M":10,"N":9})
freeze_and_filter(ws4, "C12", f"A{row_off4+1}:N{row_off4+1+len(summary_df_sorted)}")

# ════════════════════════════════════════════════════════════════════════
# SHEET 5: Best Model × Best Window
# ════════════════════════════════════════════════════════════════════════
print("Sheet 5: Best Model x Window...")
ws5 = wb.create_sheet("5. Best Model x Window")
ws5.sheet_view.showGridLines = False

ws5.merge_cells("A1:J1")
ws5["A1"] = "BEST MODEL × BEST WINDOW PER STOCK — 60% TARGET TRACKING"
ws5["A1"].font = mk_font(bold=True, size=13, color="FFFFFF")
ws5["A1"].fill = HEADER_FILL
ws5["A1"].alignment = mk_align("center", "center")
ws5.row_dimensions[1].height = 25

# Stats box
ws5.merge_cells("A2:J2")
ws5["A2"] = (f"Target: 60% | Achieved by {n_hit_60_bestmw}/100 stocks with optimal model/window | "
             f"Average Best M/W Accuracy: {overall_best_mw*100:.2f}%")
ws5["A2"].font = mk_font(bold=True, size=11, color="1F3864")
ws5["A2"].alignment = mk_align("center", "center")
ws5["A2"].fill = hex_fill("D6E4F0")

headers5 = ["Rank","Symbol","Ensemble\nOOS Acc","Best\nWindow",
            "Best\nModel","Best M/W\nAccuracy","Gap to\n60% Target",
            "Status","Dir Acc UP","Dir Acc DN"]
for c, h in enumerate(headers5, 1):
    ws5.cell(row=3, column=c).value = h
style_header_row(ws5, 3, len(headers5))
ws5.row_dimensions[3].height = 35

# Sort by best M/W accuracy
sym_best_list = []
for sym in summary_df_sorted["symbol"]:
    info = sym_best.get(sym, (None, None, 0))
    row = summary_df[summary_df["symbol"]==sym].iloc[0]
    sym_best_list.append({
        "symbol": sym,
        "oos_acc": row["oos_accuracy"],
        "best_win": info[0],
        "best_model_col": info[1],
        "best_acc": info[2] or 0,
        "dir_up": row.get("avg_dir_acc_up", 0),
        "dir_dn": row.get("avg_dir_acc_down", 0),
    })
sym_best_list.sort(key=lambda x: x["best_acc"], reverse=True)

model_col_to_name = dict(zip(MODEL_COLS, MODEL_NAMES))

for i, d in enumerate(sym_best_list, 1):
    r = i + 3
    gap = d["best_acc"] - 0.60
    if d["best_acc"] >= 0.65:
        status = "EXCELLENT"
        sf = GREEN_FILL
    elif d["best_acc"] >= 0.60:
        status = "TARGET MET"
        sf = hex_fill("C6EFCE")
    elif d["best_acc"] >= 0.55:
        status = "CLOSE"
        sf = YELLOW_FILL
    elif d["best_acc"] >= 0.50:
        status = "MARGINAL"
        sf = ORANGE_FILL
    else:
        status = "BELOW TARGET"
        sf = RED_FILL

    vals = [
        i, d["symbol"],
        d["oos_acc"], f"Win {d['best_win']}",
        model_col_to_name.get(d["best_model_col"], d["best_model_col"]),
        d["best_acc"], gap,
        status, d["dir_up"], d["dir_dn"]
    ]
    for c, v in enumerate(vals, 1):
        cell = ws5.cell(row=r, column=c)
        if isinstance(v, float):
            cell.value = round(v, 4)
            cell.number_format = "0.00%"
            if c in (3, 6, 9, 10):
                cell.fill = color_for_acc(v)
            elif c == 7:
                cell.fill = GREEN_FILL if v >= 0 else RED_FILL
            else:
                cell.fill = sf
        else:
            cell.value = v
            cell.fill = sf if c == 8 else (ALT_FILL if i%2==0 else WHITE_FILL)
        cell.font = mk_font(bold=(c==2), size=10)
        cell.alignment = mk_align("center", "center")
        cell.border = THIN_BORDER
    ws5.row_dimensions[r].height = 17

set_col_width(ws5, {"A":6,"B":13,"C":13,"D":11,"E":16,"F":13,"G":12,"H":13,"I":12,"J":12})
freeze_and_filter(ws5, "C4", f"A3:J{len(sym_best_list)+3}")

# ════════════════════════════════════════════════════════════════════════
# SHEET 6: Next Day Signals
# ════════════════════════════════════════════════════════════════════════
print("Sheet 6: Next Day Signals...")
ws6 = wb.create_sheet("6. Next Day Signals")
ws6.sheet_view.showGridLines = False

ws6.merge_cells("A1:K1")
ws6["A1"] = "NEXT-DAY TRADING SIGNALS — PRODUCTION MODEL OUTPUT"
ws6["A1"].font = mk_font(bold=True, size=13, color="FFFFFF")
ws6["A1"].fill = HEADER_FILL
ws6["A1"].alignment = mk_align("center", "center")
ws6.row_dimensions[1].height = 25

buy_df  = preds_df[preds_df["action"]=="BUY"].sort_values("confidence", ascending=False)
sell_df = preds_df[preds_df["action"]=="SELL"].sort_values("confidence", ascending=False)
hold_df = preds_df[~preds_df["action"].isin(["BUY","SELL"])]

headers6 = ["#","Symbol","Last Date","Last Close","Direction","Action",
            "Confidence","Avg Prob","Signal Active","Regime","Temperature"]
for c, h in enumerate(headers6, 1):
    ws6.cell(row=2, column=c).value = h
style_header_row(ws6, 2, len(headers6))
ws6.row_dimensions[2].height = 22

all_sigs = pd.concat([buy_df, sell_df, hold_df]).reset_index(drop=True)
for i, row in all_sigs.iterrows():
    r = i + 3
    action = row.get("action", "HOLD")
    if action == "BUY":
        af = GREEN_FILL
    elif action == "SELL":
        af = RED_FILL
    else:
        af = YELLOW_FILL

    vals = [
        i+1, row.get("symbol",""), row.get("last_date",""),
        round(float(row.get("last_close",0)), 2),
        row.get("direction",""), action,
        round(float(row.get("confidence",0)), 4),
        round(float(row.get("avg_prob",0)), 4),
        str(row.get("signal_active","")),
        row.get("regime_label",""), row.get("temperature","")
    ]
    for c, v in enumerate(vals, 1):
        cell = ws6.cell(row=r, column=c)
        cell.value = v
        if c == 6:
            cell.fill = af
            cell.font = mk_font(bold=True, size=10, color=("1F6B2E" if action=="BUY" else ("9C0006" if action=="SELL" else "000000")))
        elif c in (7, 8):
            conf_val = float(v) if isinstance(v, (int, float)) else 0
            cell.fill = GREEN_FILL if conf_val >= 0.65 else (YELLOW_FILL if conf_val >= 0.55 else WHITE_FILL)
            cell.font = mk_font(size=10)
            cell.number_format = "0.00%"
        else:
            cell.fill = ALT_FILL if i%2==0 else WHITE_FILL
            cell.font = mk_font(size=10)
        cell.alignment = mk_align("center", "center")
        cell.border = THIN_BORDER
    ws6.row_dimensions[r].height = 17

set_col_width(ws6, {"A":5,"B":13,"C":14,"D":13,"E":12,"F":11,"G":12,"H":11,"I":13,"J":12,"K":13})
freeze_and_filter(ws6, "B3", f"A2:K{len(all_sigs)+2}")

# ════════════════════════════════════════════════════════════════════════
# SHEET 7: Attention Required
# ════════════════════════════════════════════════════════════════════════
print("Sheet 7: Attention Required...")
ws7 = wb.create_sheet("7. Attention Required")
ws7.sheet_view.showGridLines = False

below50 = summary_df[summary_df["oos_accuracy"] < 0.50].sort_values("oos_accuracy").reset_index(drop=True)

ws7.merge_cells("A1:L1")
ws7["A1"] = f"STOCKS REQUIRING ATTENTION — {len(below50)} STOCKS BELOW 50% OOS ACCURACY"
ws7["A1"].font = mk_font(bold=True, size=13, color="FFFFFF")
ws7["A1"].fill = hex_fill("C00000")
ws7["A1"].alignment = mk_align("center", "center")
ws7.row_dimensions[1].height = 25

ws7.merge_cells("A2:L2")
ws7["A2"] = "These stocks perform below random baseline. They may need longer data history, sector-specific features, or exclusion from trading universe."
ws7["A2"].font = mk_font(italic=True, size=10, color="9C0006")
ws7["A2"].alignment = mk_align("center", "center", wrap=True)
ws7["A2"].fill = hex_fill("FCE4D6")

headers7 = ["Rank","Symbol","Ensemble\nOOS Acc","Best M/W\nAcc","Best Model","Best Window",
            "Gap to 50%","Gap to 60%","Recommended\nAction","Dir UP","Dir DN","Avg F1"]
for c, h in enumerate(headers7, 1):
    ws7.cell(row=3, column=c).value = h
style_header_row(ws7, 3, len(headers7), fill=hex_fill("C00000"))
ws7.row_dimensions[3].height = 35

for i, row in below50.iterrows():
    r = i + 4
    best_info = sym_best.get(row["symbol"], (None, None, 0))
    best_mn = model_col_to_name.get(best_info[1], "") if best_info[1] else ""
    ens_acc = row["oos_accuracy"]
    best_acc = best_info[2] or 0
    gap50 = ens_acc - 0.50
    gap60 = ens_acc - 0.60

    if ens_acc < 0.40:
        action_rec = "Exclude from universe"
        af = hex_fill("C00000")
    elif ens_acc < 0.45:
        action_rec = "Retrain with more data"
        af = RED_FILL
    else:
        action_rec = "Add sector features / tune"
        af = ORANGE_FILL

    vals = [
        i+1, row["symbol"], ens_acc, best_acc, best_mn,
        f"Win {best_info[0]}" if best_info[0] else "N/A",
        gap50, gap60, action_rec,
        row.get("avg_dir_acc_up",0), row.get("avg_dir_acc_down",0), row.get("oos_f1",0)
    ]
    for c, v in enumerate(vals, 1):
        cell = ws7.cell(row=r, column=c)
        if isinstance(v, float):
            cell.value = round(v, 4)
            cell.number_format = "0.00%"
            cell.fill = af if c in (3,) else (color_for_acc(v) if c in (4,10,11,12) else af)
        else:
            cell.value = v
            cell.fill = af if c != 9 else (hex_fill("C00000") if action_rec.startswith("Exclude") else (RED_FILL if action_rec.startswith("Retrain") else ORANGE_FILL))
        cell.font = mk_font(bold=(c==2), size=10,
                            color=("FFFFFF" if af == hex_fill("C00000") else "000000"))
        cell.alignment = mk_align("center", "center", wrap=(c==9))
        cell.border = THIN_BORDER
    ws7.row_dimensions[r].height = 20

set_col_width(ws7, {"A":6,"B":13,"C":13,"D":12,"E":16,"F":12,"G":11,"H":11,
                    "I":25,"J":11,"K":11,"L":11})
freeze_and_filter(ws7, "C4")

# ════════════════════════════════════════════════════════════════════════
# SHEET 8: Raw Data
# ════════════════════════════════════════════════════════════════════════
print("Sheet 8: Raw Data...")
ws8 = wb.create_sheet("8. Raw Window Data")
ws8.sheet_view.showGridLines = False

ws8.merge_cells("A1:AH1")
ws8["A1"] = "RAW WINDOW-LEVEL DATA — ALL 100 STOCKS × 6 WINDOWS (600 rows)"
ws8["A1"].font = mk_font(bold=True, size=12, color="FFFFFF")
ws8["A1"].fill = HEADER_FILL
ws8["A1"].alignment = mk_align("center", "center")
ws8.row_dimensions[1].height = 22

cols8 = list(detail_df.columns)
for c, h in enumerate(cols8, 1):
    cell = ws8.cell(row=2, column=c)
    cell.value = h
    cell.fill = SUBHEAD_FILL
    cell.font = mk_font(bold=True, size=9, color="FFFFFF")
    cell.alignment = mk_align("center", "center", wrap=True)
    cell.border = THIN_BORDER
ws8.row_dimensions[2].height = 30

for i, row in detail_df.iterrows():
    r = i + 3
    for c, v in enumerate(row, 1):
        cell = ws8.cell(row=r, column=c)
        if isinstance(v, float):
            cell.value = round(v, 6)
            if c > 8:
                cell.number_format = "0.0000"
        elif pd.isna(v):
            cell.value = ""
        else:
            cell.value = v
        cell.fill = ALT_FILL if i%2==0 else WHITE_FILL
        cell.font = mk_font(size=8)
        cell.alignment = mk_align("center", "center")
        cell.border = THIN_BORDER
    ws8.row_dimensions[r].height = 13

for c in range(1, len(cols8)+1):
    ws8.column_dimensions[get_column_letter(c)].width = 12
ws8.column_dimensions["A"].width = 14
ws8.column_dimensions["B"].width = 8
freeze_and_filter(ws8, "C3", f"A2:{get_column_letter(len(cols8))}{len(detail_df)+2}")

# ════════════════════════════════════════════════════════════════════════
# Tab colors
# ════════════════════════════════════════════════════════════════════════
ws1.sheet_properties.tabColor = "1F3864"
ws2.sheet_properties.tabColor = "2E75B6"
ws3.sheet_properties.tabColor = "375623"
ws4.sheet_properties.tabColor = "7030A0"
ws5.sheet_properties.tabColor = "C55A11"
ws6.sheet_properties.tabColor = "538135"
ws7.sheet_properties.tabColor = "C00000"
ws8.sheet_properties.tabColor = "595959"

# ════════════════════════════════════════════════════════════════════════
# Save
# ════════════════════════════════════════════════════════════════════════
print(f"Saving to {OUT_FILE} ...")
wb.save(OUT_FILE)
print(f"OK: Saved {OUT_FILE}")
print(f"File size: {os.path.getsize(OUT_FILE) / 1024:.1f} KB")
