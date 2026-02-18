"""
V1 vs V2 Pipeline Comparison - Excel Report Generator
"""
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

RESULTS_DIR = os.path.join(os.path.dirname(__file__), 'results')

def load_data():
    v1 = pd.read_csv(os.path.join(RESULTS_DIR, 'all_stocks_summary.csv'))
    v2 = pd.read_csv(os.path.join(RESULTS_DIR, 'v2_summary.csv'))
    return v1, v2

def style_header(ws, row, max_col):
    hdr_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    hdr_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

def acc_fill(val):
    if val >= 0.58: return PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid')
    if val >= 0.54: return PatternFill(start_color='82E0AA', end_color='82E0AA', fill_type='solid')
    if val >= 0.52: return PatternFill(start_color='F9E79F', end_color='F9E79F', fill_type='solid')
    if val >= 0.50: return PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')
    return PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid')

def delta_fill(val):
    if val > 0.03: return PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid')
    if val > 0.01: return PatternFill(start_color='82E0AA', end_color='82E0AA', fill_type='solid')
    if val > 0: return PatternFill(start_color='F9E79F', end_color='F9E79F', fill_type='solid')
    if val > -0.02: return PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')
    return PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid')

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 25)

def create_executive_summary(wb, v1, v2):
    ws = wb.active
    ws.title = 'Executive Summary'

    # V1 ensemble stats
    v1_ens = v1[v1['model'] == 'ensemble']
    v2_ens = v2[v2['model'] == 'ensemble']

    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = 'V1 vs V2 Pipeline Comparison - Executive Summary'
    ws['A1'].font = Font(name='Calibri', bold=True, size=16, color='1F4E79')

    ws.merge_cells('A2:H2')
    ws['A2'] = 'V1: Regression (XGB+LGB) | V2: Classification (XGB+LGB+CatBoost) + Feature Selection + Ensemble Fix'
    ws['A2'].font = Font(name='Calibri', size=11, italic=True)

    # Overall comparison table
    row = 4
    headers = ['Metric', 'V1 (Regression)', 'V2 (Classification)', 'Delta', 'Change %', 'Verdict']
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header(ws, row, len(headers))

    metrics = [
        ('Avg Dir Accuracy (Ensemble)', v1_ens['dir_accuracy'].mean(), v2_ens['dir_accuracy'].mean()),
        ('Best Dir Accuracy', v1_ens['dir_accuracy'].max(), v2_ens['dir_accuracy'].max()),
        ('Avg Win Rate Long', v1_ens['win_rate_long'].mean(), v2_ens['win_rate_long'].mean()),
        ('Avg Win Rate Short', v1_ens['win_rate_short'].mean(), v2_ens['win_rate_short'].mean()),
        ('Avg Sharpe Ratio', v1_ens['sharpe'].mean(), v2_ens['sharpe'].mean()),
        ('Avg Profit Factor', v1_ens['profit_factor'].mean(), v2_ens['profit_factor'].mean()),
        ('Stocks with Acc > 55%', (v1_ens.groupby('symbol')['dir_accuracy'].max() > 0.55).sum(),
         (v2_ens.groupby('symbol')['dir_accuracy'].max() > 0.55).sum()),
        ('Stocks with Acc > 58%', (v1_ens.groupby('symbol')['dir_accuracy'].max() > 0.58).sum(),
         (v2_ens.groupby('symbol')['dir_accuracy'].max() > 0.58).sum()),
    ]

    for i, (name, v1_val, v2_val) in enumerate(metrics):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=name)

        if isinstance(v1_val, (int, np.integer)):
            ws.cell(row=r, column=2, value=v1_val)
            ws.cell(row=r, column=3, value=v2_val)
            delta = v2_val - v1_val
            ws.cell(row=r, column=4, value=delta)
            ws.cell(row=r, column=5, value=f"{delta:+d}")
        else:
            ws.cell(row=r, column=2, value=f"{v1_val:.4f}")
            ws.cell(row=r, column=3, value=f"{v2_val:.4f}")
            delta = v2_val - v1_val
            ws.cell(row=r, column=4, value=f"{delta:+.4f}")
            pct = (delta / abs(v1_val) * 100) if v1_val != 0 else 0
            ws.cell(row=r, column=5, value=f"{pct:+.1f}%")

        verdict = 'IMPROVED' if (v2_val > v1_val if not isinstance(v1_val, (int, np.integer)) else v2_val > v1_val) else ('SAME' if v2_val == v1_val else 'WORSE')
        ws.cell(row=r, column=6, value=verdict)
        if verdict == 'IMPROVED':
            ws.cell(row=r, column=6).font = Font(color='27AE60', bold=True)
        elif verdict == 'WORSE':
            ws.cell(row=r, column=6).font = Font(color='E74C3C', bold=True)

    # Per-stock best accuracy comparison
    row = row + len(metrics) + 3
    ws.cell(row=row, column=1, value='Per-Stock Best Accuracy (Ensemble)')
    ws.cell(row=row, column=1).font = Font(bold=True, size=13, color='1F4E79')
    row += 1

    headers2 = ['Stock', 'Sector', 'V1 Best Acc', 'V1 Best Window', 'V2 Best Acc', 'V2 Best Window', 'Delta', 'V2 Best Sharpe', 'V2 Best PF']
    for c, h in enumerate(headers2, 1):
        ws.cell(row=row, column=c, value=h)
    style_header(ws, row, len(headers2))

    sectors = {'HDFCBANK': 'Banking', 'ICICIBANK': 'Banking', 'KOTAKBANK': 'Banking', 'SBIN': 'Banking', 'AXISBANK': 'Banking',
               'TCS': 'IT', 'INFY': 'IT', 'HCLTECH': 'IT', 'WIPRO': 'IT', 'TECHM': 'IT'}

    for sym in sorted(v2_ens['symbol'].unique()):
        row += 1
        # V1 best
        v1_sym = v1_ens[v1_ens['symbol'] == sym]
        if len(v1_sym) > 0:
            v1_best_idx = v1_sym['dir_accuracy'].idxmax()
            v1_best_acc = v1_sym.loc[v1_best_idx, 'dir_accuracy']
            v1_best_win = v1_sym.loc[v1_best_idx, 'window']
        else:
            v1_best_acc = 0
            v1_best_win = 'N/A'

        # V2 best
        v2_sym = v2_ens[v2_ens['symbol'] == sym]
        v2_best_idx = v2_sym['dir_accuracy'].idxmax()
        v2_best_acc = v2_sym.loc[v2_best_idx, 'dir_accuracy']
        v2_best_win = v2_sym.loc[v2_best_idx, 'window']
        v2_best_sharpe = v2_sym.loc[v2_best_idx, 'sharpe']
        v2_best_pf = v2_sym.loc[v2_best_idx, 'profit_factor']

        delta = v2_best_acc - v1_best_acc

        ws.cell(row=row, column=1, value=sym)
        ws.cell(row=row, column=2, value=sectors.get(sym, ''))
        ws.cell(row=row, column=3, value=f"{v1_best_acc:.1%}")
        ws.cell(row=row, column=4, value=v1_best_win)
        ws.cell(row=row, column=5, value=f"{v2_best_acc:.1%}")
        ws.cell(row=row, column=6, value=v2_best_win)
        ws.cell(row=row, column=7, value=f"{delta:+.1%}")
        ws.cell(row=row, column=8, value=f"{v2_best_sharpe:.2f}")
        ws.cell(row=row, column=9, value=f"{v2_best_pf:.2f}")

        ws.cell(row=row, column=5).fill = acc_fill(v2_best_acc)
        ws.cell(row=row, column=7).fill = delta_fill(delta)

    # Key changes section
    row += 3
    ws.cell(row=row, column=1, value='Key Changes V1 -> V2')
    ws.cell(row=row, column=1).font = Font(bold=True, size=13, color='1F4E79')

    changes = [
        ('Models', 'XGBRegressor + LGBMRegressor', 'XGBClassifier + LGBMClassifier + CatBoostClassifier'),
        ('Target', 'Continuous return (regression)', 'Binary UP/DOWN (classification)'),
        ('Features', '240 (all)', '50 (mutual_info_classif selection)'),
        ('Ensemble', 'scipy.optimize (stuck at 0.5/0.5)', 'Grid search 9 weight combos (3 models)'),
        ('Confidence', 'None', '0.55 threshold filter'),
        ('Windows', '7 (including tiny 96/3/1)', '5 (min 70/15/15)'),
        ('Early Stopping', 'On test set (leakage)', 'On separate val set'),
    ]
    for i, (aspect, v1_val, v2_val) in enumerate(changes):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=aspect).font = Font(bold=True)
        ws.cell(row=r, column=2, value=v1_val)
        ws.cell(row=r, column=4, value=v2_val)

    auto_width(ws)

def create_per_stock_detail(wb, v2):
    ws = wb.create_sheet('Per Stock V2 Detail')

    headers = ['Symbol', 'Window', 'Model', 'Features', 'Dir Acc', 'WR Long', 'WR Short',
               'N Long', 'N Short', 'Sharpe', 'PF', 'Filt Acc', 'Filt N', 'Filter Rate',
               'W_XGB', 'W_LGB', 'W_CAT']

    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))

    for i, (_, r) in enumerate(v2.iterrows()):
        row = i + 2
        ws.cell(row=row, column=1, value=r['symbol'])
        ws.cell(row=row, column=2, value=r['window'])
        ws.cell(row=row, column=3, value=r['model'])
        ws.cell(row=row, column=4, value=r['n_features'])
        ws.cell(row=row, column=5, value=f"{r['dir_accuracy']:.1%}")
        ws.cell(row=row, column=6, value=f"{r['win_rate_long']:.1%}")
        ws.cell(row=row, column=7, value=f"{r['win_rate_short']:.1%}")
        ws.cell(row=row, column=8, value=r['n_long'])
        ws.cell(row=row, column=9, value=r['n_short'])
        ws.cell(row=row, column=10, value=f"{r['sharpe']:.2f}")
        ws.cell(row=row, column=11, value=f"{r['profit_factor']:.2f}")
        ws.cell(row=row, column=12, value=f"{r['filtered_accuracy']:.1%}" if pd.notna(r.get('filtered_accuracy')) else 'N/A')
        ws.cell(row=row, column=13, value=r.get('n_filtered_trades', 0))
        ws.cell(row=row, column=14, value=f"{r.get('filter_rate', 0):.1%}")
        ws.cell(row=row, column=15, value=r.get('w_xgb', ''))
        ws.cell(row=row, column=16, value=r.get('w_lgb', ''))
        ws.cell(row=row, column=17, value=r.get('w_cat', ''))

        # Color accuracy
        ws.cell(row=row, column=5).fill = acc_fill(r['dir_accuracy'])

        # Highlight ensemble rows
        if r['model'] == 'ensemble':
            for c in range(1, len(headers)+1):
                ws.cell(row=row, column=c).font = Font(bold=True)

    auto_width(ws)

def create_model_comparison(wb, v2):
    ws = wb.create_sheet('Model Comparison V2')

    ws.merge_cells('A1:F1')
    ws['A1'] = 'Model Performance Comparison (V2 Pipeline)'
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79')

    # Per model average
    row = 3
    headers = ['Model', 'Avg Acc', 'Avg WR Long', 'Avg WR Short', 'Avg Sharpe', 'Avg PF', 'Best Acc', 'Worst Acc']
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header(ws, row, len(headers))

    for model in ['xgb', 'lgb', 'catboost', 'ensemble']:
        row += 1
        m = v2[v2['model'] == model]
        ws.cell(row=row, column=1, value=model.upper())
        ws.cell(row=row, column=2, value=f"{m['dir_accuracy'].mean():.1%}")
        ws.cell(row=row, column=3, value=f"{m['win_rate_long'].mean():.1%}")
        ws.cell(row=row, column=4, value=f"{m['win_rate_short'].mean():.1%}")
        ws.cell(row=row, column=5, value=f"{m['sharpe'].mean():.2f}")
        ws.cell(row=row, column=6, value=f"{m['profit_factor'].mean():.2f}")
        ws.cell(row=row, column=7, value=f"{m['dir_accuracy'].max():.1%}")
        ws.cell(row=row, column=8, value=f"{m['dir_accuracy'].min():.1%}")
        ws.cell(row=row, column=2).fill = acc_fill(m['dir_accuracy'].mean())

    # Best model per stock
    row += 3
    ws.cell(row=row, column=1, value='Best Model Per Stock (highest accuracy across all windows)')
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color='1F4E79')
    row += 1

    headers2 = ['Stock', 'Best Model', 'Best Window', 'Accuracy', 'Sharpe', 'PF', 'WR Long', 'WR Short']
    for c, h in enumerate(headers2, 1):
        ws.cell(row=row, column=c, value=h)
    style_header(ws, row, len(headers2))

    for sym in sorted(v2['symbol'].unique()):
        row += 1
        sym_data = v2[v2['symbol'] == sym]
        best_idx = sym_data['dir_accuracy'].idxmax()
        best = sym_data.loc[best_idx]

        ws.cell(row=row, column=1, value=sym)
        ws.cell(row=row, column=2, value=best['model'].upper())
        ws.cell(row=row, column=3, value=best['window'])
        ws.cell(row=row, column=4, value=f"{best['dir_accuracy']:.1%}")
        ws.cell(row=row, column=5, value=f"{best['sharpe']:.2f}")
        ws.cell(row=row, column=6, value=f"{best['profit_factor']:.2f}")
        ws.cell(row=row, column=7, value=f"{best['win_rate_long']:.1%}")
        ws.cell(row=row, column=8, value=f"{best['win_rate_short']:.1%}")
        ws.cell(row=row, column=4).fill = acc_fill(best['dir_accuracy'])

    # Ensemble weight distribution
    row += 3
    ws.cell(row=row, column=1, value='Ensemble Weight Distribution (V2)')
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color='1F4E79')
    row += 1

    ens = v2[v2['model'] == 'ensemble']
    weight_counts = ens.groupby(['w_xgb', 'w_lgb', 'w_cat']).size().reset_index(name='count')
    weight_counts = weight_counts.sort_values('count', ascending=False)

    headers3 = ['W_XGB', 'W_LGB', 'W_CAT', 'Count', 'Avg Accuracy']
    for c, h in enumerate(headers3, 1):
        ws.cell(row=row, column=c, value=h)
    style_header(ws, row, len(headers3))

    for _, wc in weight_counts.iterrows():
        row += 1
        mask = (ens['w_xgb'] == wc['w_xgb']) & (ens['w_lgb'] == wc['w_lgb']) & (ens['w_cat'] == wc['w_cat'])
        avg_acc = ens[mask]['dir_accuracy'].mean()
        ws.cell(row=row, column=1, value=wc['w_xgb'])
        ws.cell(row=row, column=2, value=wc['w_lgb'])
        ws.cell(row=row, column=3, value=wc['w_cat'])
        ws.cell(row=row, column=4, value=wc['count'])
        ws.cell(row=row, column=5, value=f"{avg_acc:.1%}")

    auto_width(ws)

def create_sector_analysis(wb, v2):
    ws = wb.create_sheet('Sector Analysis')

    sectors = {'HDFCBANK': 'Banking', 'ICICIBANK': 'Banking', 'KOTAKBANK': 'Banking', 'SBIN': 'Banking', 'AXISBANK': 'Banking',
               'TCS': 'IT', 'INFY': 'IT', 'HCLTECH': 'IT', 'WIPRO': 'IT', 'TECHM': 'IT'}

    v2_ens = v2[v2['model'] == 'ensemble'].copy()
    v2_ens['sector'] = v2_ens['symbol'].map(sectors)

    ws.merge_cells('A1:F1')
    ws['A1'] = 'Sector Performance Analysis (V2 Ensemble)'
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79')

    row = 3
    headers = ['Sector', 'Avg Acc', 'Best Acc', 'Avg Sharpe', 'Avg PF', 'Best Stock']
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header(ws, row, len(headers))

    for sector in ['Banking', 'IT']:
        row += 1
        sec = v2_ens[v2_ens['sector'] == sector]
        best_stock = sec.loc[sec['dir_accuracy'].idxmax(), 'symbol']
        ws.cell(row=row, column=1, value=sector)
        ws.cell(row=row, column=2, value=f"{sec['dir_accuracy'].mean():.1%}")
        ws.cell(row=row, column=3, value=f"{sec['dir_accuracy'].max():.1%}")
        ws.cell(row=row, column=4, value=f"{sec['sharpe'].mean():.2f}")
        ws.cell(row=row, column=5, value=f"{sec['profit_factor'].mean():.2f}")
        ws.cell(row=row, column=6, value=best_stock)
        ws.cell(row=row, column=2).fill = acc_fill(sec['dir_accuracy'].mean())

    auto_width(ws)

def create_next_steps(wb, v1, v2):
    ws = wb.create_sheet('Next Steps')

    v1_ens = v1[v1['model'] == 'ensemble']
    v2_ens = v2[v2['model'] == 'ensemble']

    ws.merge_cells('A1:D1')
    ws['A1'] = 'Analysis & Next Steps for Accuracy Improvement'
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79')

    # Current status
    row = 3
    ws.cell(row=row, column=1, value='CURRENT STATUS').font = Font(bold=True, size=12, color='E74C3C')
    row += 1

    status_items = [
        f"V2 avg ensemble accuracy: {v2_ens['dir_accuracy'].mean():.1%} (V1 was {v1_ens['dir_accuracy'].mean():.1%})",
        f"Improvement: {(v2_ens['dir_accuracy'].mean() - v1_ens['dir_accuracy'].mean()):.1%} absolute",
        f"V2 best single result: {v2_ens['dir_accuracy'].max():.1%}",
        f"Target: 58%+ average, 65%+ best",
        f"Gap to target: {0.58 - v2_ens['dir_accuracy'].mean():.1%}",
    ]
    for s in status_items:
        ws.cell(row=row, column=1, value=s)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value='REMAINING ISSUES').font = Font(bold=True, size=12, color='E74C3C')
    row += 1

    issues = [
        ('1. Features still too noisy', 'Top mutual_info features have weak signal (max MI < 0.05). Need domain-specific features.'),
        ('2. No temporal modeling', 'Tree models treat each row independently. Stock prices have autocorrelation and momentum.'),
        ('3. No regime awareness', 'Same model for trending vs mean-reverting markets. Need regime-specific models.'),
        ('4. CatBoost underperforming', 'CatBoost often has lowest accuracy. Needs better hyperparameter tuning.'),
        ('5. Filtered accuracy not helping', 'Confidence filtering sometimes REDUCES accuracy - threshold too conservative or not calibrated.'),
    ]
    for title, desc in issues:
        ws.cell(row=row, column=1, value=title).font = Font(bold=True)
        ws.cell(row=row, column=2, value=desc)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value='IMPROVEMENT ROADMAP').font = Font(bold=True, size=12, color='27AE60')
    row += 1

    headers = ['Priority', 'Change', 'Expected Impact', 'Effort', 'Details']
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header(ws, row, len(headers))

    improvements = [
        ('P0', 'Add LSTM/Neural models', '+5-10% accuracy', 'High', 'Already implemented in V3/neural_models.py. Sequence models capture temporal patterns that tree models miss.'),
        ('P0', 'Hyperparameter tuning (Optuna)', '+2-4% accuracy', 'Medium', 'Use Optuna for XGB/LGB/CatBoost hyperparameter optimization per stock.'),
        ('P1', 'Feature engineering v2', '+3-5% accuracy', 'Medium', 'Add interaction features, rolling z-scores, volatility-adjusted returns, order flow imbalance.'),
        ('P1', 'Walk-forward validation', '+1-3% accuracy', 'Medium', 'Instead of single split, use rolling 252-day train, 63-day test windows.'),
        ('P1', 'Calibrated confidence', '+2-3% win rate', 'Low', 'Use Platt scaling or isotonic regression for probability calibration.'),
        ('P2', 'Regime-aware models', '+3-5% accuracy', 'High', 'Train separate models for bull/bear/sideways markets. Use HMM for regime detection.'),
        ('P2', 'Stacking ensemble', '+2-4% accuracy', 'Medium', 'Use meta-learner (logistic regression) on base model predictions instead of weighted average.'),
        ('P3', 'Transformer model', '+3-8% accuracy', 'High', 'Attention mechanism for multi-scale temporal patterns. Already in V3/neural_models.py.'),
    ]

    for p, change, impact, effort, details in improvements:
        row += 1
        ws.cell(row=row, column=1, value=p)
        ws.cell(row=row, column=2, value=change)
        ws.cell(row=row, column=3, value=impact)
        ws.cell(row=row, column=4, value=effort)
        ws.cell(row=row, column=5, value=details)

        if p == 'P0':
            ws.cell(row=row, column=1).fill = PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid')
            ws.cell(row=row, column=1).font = Font(color='FFFFFF', bold=True)
        elif p == 'P1':
            ws.cell(row=row, column=1).fill = PatternFill(start_color='F39C12', end_color='F39C12', fill_type='solid')
            ws.cell(row=row, column=1).font = Font(bold=True)

    auto_width(ws)

def create_top_features(wb, v2):
    """Show which features are most commonly selected across stocks/windows"""
    ws = wb.create_sheet('Top Features')

    ws.merge_cells('A1:D1')
    ws['A1'] = 'Most Important Features (by frequency of selection across all runs)'
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79')

    # We can't extract feature importance from CSV, but we can note the features from the logs
    row = 3
    ws.cell(row=row, column=1, value='Note: Feature importance tracked from pipeline logs during V2 run.').font = Font(italic=True)

    row = 5
    headers = ['Feature', 'Category', 'Appeared In Top 5', 'Notes']
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header(ws, row, len(headers))

    # From the pipeline output, these were the most common top-5 features:
    features = [
        ('price_zscore', 'Statistical', 'Frequent', 'Mean reversion signal - distance from mean'),
        ('parkinson_vol', 'Volatility', 'Frequent', 'Parkinson volatility estimator (range-based)'),
        ('ultimate_oscillator', 'Momentum', 'Frequent', 'Multi-timeframe oscillator'),
        ('supertrend', 'Technical', 'Frequent', 'Trend following indicator'),
        ('regime_changes_30d', 'Regime', 'Frequent', 'Market regime shift detection'),
        ('kst_histogram', 'Momentum', 'Frequent', 'Know Sure Thing oscillator'),
        ('cci_20', 'Momentum', 'Frequent', 'Commodity Channel Index'),
        ('is_doji', 'Pattern', 'Frequent', 'Candlestick doji pattern'),
        ('return_1d', 'Price', 'Common', '1-day return'),
        ('atr_7', 'Volatility', 'Common', '7-day Average True Range'),
        ('vol_change_5d', 'Volume', 'Common', '5-day volume change'),
        ('bb_width', 'Technical', 'Common', 'Bollinger Band width'),
        ('mfi_20', 'Volume', 'Common', 'Money Flow Index 20'),
        ('dpo', 'Technical', 'Common', 'Detrended Price Oscillator'),
        ('momentum_rank_10d', 'Momentum', 'Common', '10-day momentum rank'),
    ]

    for feat, cat, freq, notes in features:
        row += 1
        ws.cell(row=row, column=1, value=feat)
        ws.cell(row=row, column=2, value=cat)
        ws.cell(row=row, column=3, value=freq)
        ws.cell(row=row, column=4, value=notes)

    auto_width(ws)

def main():
    v1, v2 = load_data()
    wb = Workbook()

    create_executive_summary(wb, v1, v2)
    create_per_stock_detail(wb, v2)
    create_model_comparison(wb, v2)
    create_sector_analysis(wb, v2)
    create_next_steps(wb, v1, v2)
    create_top_features(wb, v2)

    out_path = os.path.join(RESULTS_DIR, 'V1_vs_V2_Comparison.xlsx')
    wb.save(out_path)
    print(f"Saved: {out_path}")
    print(f"V1 rows: {len(v1)}, V2 rows: {len(v2)}")

if __name__ == '__main__':
    main()
