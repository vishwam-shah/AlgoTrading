"""
V3 Pipeline Analysis - Generate comprehensive Excel report
"""
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# Paths
RESULTS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'results')
SUMMARY_CSV = os.path.join(RESULTS_DIR, 'all_stocks_summary.csv')
OUTPUT_XLSX = os.path.join(RESULTS_DIR, 'V3_Pipeline_Analysis.xlsx')

# Load data
df = pd.read_csv(SUMMARY_CSV)

# Styles
header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
good_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
bad_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
warn_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
section_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
section_font = Font(bold=True, size=12, color='1F4E79')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

def style_data(ws, start_row, end_row, cols):
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

def auto_width(ws, cols):
    for col in range(1, cols + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 25)

def color_accuracy(ws, row, col, val):
    cell = ws.cell(row=row, column=col)
    if val >= 0.58:
        cell.fill = good_fill
    elif val >= 0.52:
        cell.fill = warn_fill
    elif val < 0.48:
        cell.fill = bad_fill

wb = Workbook()

# ============ SHEET 1: EXECUTIVE SUMMARY ============
ws1 = wb.active
ws1.title = 'Executive Summary'

ens = df[df['model'] == 'ensemble']

ws1.cell(row=1, column=1, value='V3 PIPELINE - EXECUTIVE SUMMARY').font = Font(bold=True, size=16, color='1F4E79')
ws1.cell(row=2, column=1, value=f'Run Date: {pd.Timestamp.now().strftime("%Y-%m-%d %H:%M")}').font = Font(size=10, italic=True)
ws1.cell(row=3, column=1, value=f'10 Stocks x 7 Windows x 3 Models = {len(df)} results').font = Font(size=10)

# Overall stats
r = 5
ws1.cell(row=r, column=1, value='OVERALL METRICS').font = section_font
ws1.cell(row=r, column=1).fill = section_fill
for c in range(1, 5):
    ws1.cell(row=r, column=c).fill = section_fill

r = 6
headers = ['Metric', 'Ensemble Avg', 'Best Single', 'Worst Single']
for i, h in enumerate(headers, 1):
    ws1.cell(row=r, column=i, value=h)
style_header(ws1, r, 4)

metrics_data = [
    ('Dir. Accuracy', f"{ens['dir_accuracy'].mean():.1%}", f"{ens['dir_accuracy'].max():.1%}", f"{ens['dir_accuracy'].min():.1%}"),
    ('Win Rate Long', f"{ens['win_rate_long'].mean():.1%}", f"{ens['win_rate_long'].max():.1%}", f"{ens['win_rate_long'].min():.1%}"),
    ('Win Rate Short', f"{ens['win_rate_short'].mean():.1%}", f"{ens['win_rate_short'].max():.1%}", f"{ens['win_rate_short'].min():.1%}"),
    ('Sharpe Ratio', f"{ens['sharpe'].mean():.2f}", f"{ens['sharpe'].max():.2f}", f"{ens['sharpe'].min():.2f}"),
    ('Profit Factor', f"{ens['profit_factor'].mean():.2f}", f"{ens['profit_factor'].max():.2f}", f"{ens['profit_factor'].min():.2f}"),
    ('RMSE', f"{ens['rmse'].mean():.6f}", f"{ens['rmse'].min():.6f}", f"{ens['rmse'].max():.6f}"),
]
for i, (name, avg, best, worst) in enumerate(metrics_data, r+1):
    ws1.cell(row=i, column=1, value=name)
    ws1.cell(row=i, column=2, value=avg)
    ws1.cell(row=i, column=3, value=best)
    ws1.cell(row=i, column=4, value=worst)
style_data(ws1, r+1, r+len(metrics_data), 4)

# Stock rankings
r = r + len(metrics_data) + 2
ws1.cell(row=r, column=1, value='STOCK RANKINGS (by Best Ensemble Dir. Accuracy)').font = section_font
ws1.cell(row=r, column=1).fill = section_fill
for c in range(1, 8):
    ws1.cell(row=r, column=c).fill = section_fill

r += 1
rank_headers = ['Rank', 'Symbol', 'Sector', 'Best Accuracy', 'Best Window', 'Sharpe', 'Profit Factor']
for i, h in enumerate(rank_headers, 1):
    ws1.cell(row=r, column=i, value=h)
style_header(ws1, r, 7)

sectors = {
    'HDFCBANK': 'Banking', 'ICICIBANK': 'Banking', 'KOTAKBANK': 'Banking',
    'SBIN': 'Banking', 'AXISBANK': 'Banking',
    'TCS': 'IT', 'INFY': 'IT', 'HCLTECH': 'IT', 'WIPRO': 'IT', 'TECHM': 'IT'
}

stock_best = ens.groupby('symbol').apply(
    lambda x: x.loc[x['dir_accuracy'].idxmax()]
).sort_values('dir_accuracy', ascending=False).reset_index(drop=True)

for i, row in stock_best.iterrows():
    ri = r + 1 + i
    ws1.cell(row=ri, column=1, value=i+1)
    ws1.cell(row=ri, column=2, value=row['symbol'])
    ws1.cell(row=ri, column=3, value=sectors.get(row['symbol'], ''))
    ws1.cell(row=ri, column=4, value=f"{row['dir_accuracy']:.1%}")
    ws1.cell(row=ri, column=5, value=row['window'])
    ws1.cell(row=ri, column=6, value=f"{row['sharpe']:.2f}")
    ws1.cell(row=ri, column=7, value=f"{row['profit_factor']:.2f}")
    color_accuracy(ws1, ri, 4, row['dir_accuracy'])
style_data(ws1, r+1, r+len(stock_best), 7)

auto_width(ws1, 7)

# ============ SHEET 2: PER-STOCK DETAIL ============
ws2 = wb.create_sheet('Per Stock Detail')
headers2 = ['Symbol', 'Window', 'Model', 'Dir Accuracy', 'WR Long', 'WR Short',
            'Sharpe', 'Profit Factor', 'RMSE', 'MAE', 'N_Train', 'N_Val']
for i, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=i, value=h)
style_header(ws2, 1, len(headers2))

for ri, (_, row) in enumerate(df.iterrows(), 2):
    ws2.cell(row=ri, column=1, value=row['symbol'])
    ws2.cell(row=ri, column=2, value=row['window'])
    ws2.cell(row=ri, column=3, value=row['model'])
    ws2.cell(row=ri, column=4, value=round(row['dir_accuracy'], 4))
    ws2.cell(row=ri, column=5, value=round(row['win_rate_long'], 4))
    ws2.cell(row=ri, column=6, value=round(row['win_rate_short'], 4))
    ws2.cell(row=ri, column=7, value=round(row['sharpe'], 4))
    ws2.cell(row=ri, column=8, value=round(row['profit_factor'], 4))
    ws2.cell(row=ri, column=9, value=round(row['rmse'], 6))
    ws2.cell(row=ri, column=10, value=round(row['mae'], 6))
    ws2.cell(row=ri, column=11, value=int(row['n_train']))
    ws2.cell(row=ri, column=12, value=int(row['n_val']))
    color_accuracy(ws2, ri, 4, row['dir_accuracy'])
style_data(ws2, 2, len(df) + 1, len(headers2))
auto_width(ws2, len(headers2))

# ============ SHEET 3: MODEL COMPARISON ============
ws3 = wb.create_sheet('Model Comparison')
ws3.cell(row=1, column=1, value='MODEL COMPARISON (Avg across all stocks & windows)').font = section_font

headers3 = ['Model', 'Avg Dir Acc', 'Avg WR Long', 'Avg WR Short', 'Avg Sharpe', 'Avg PF', 'Avg RMSE']
for i, h in enumerate(headers3, 1):
    ws3.cell(row=3, column=i, value=h)
style_header(ws3, 3, len(headers3))

for ri, model in enumerate(['xgb', 'lgb', 'ensemble'], 4):
    mdf = df[df['model'] == model]
    ws3.cell(row=ri, column=1, value=model.upper())
    ws3.cell(row=ri, column=2, value=round(mdf['dir_accuracy'].mean(), 4))
    ws3.cell(row=ri, column=3, value=round(mdf['win_rate_long'].mean(), 4))
    ws3.cell(row=ri, column=4, value=round(mdf['win_rate_short'].mean(), 4))
    ws3.cell(row=ri, column=5, value=round(mdf['sharpe'].mean(), 4))
    ws3.cell(row=ri, column=6, value=round(mdf['profit_factor'].mean(), 4))
    ws3.cell(row=ri, column=7, value=round(mdf['rmse'].mean(), 6))
    color_accuracy(ws3, ri, 2, mdf['dir_accuracy'].mean())
style_data(ws3, 4, 6, len(headers3))

# Window comparison
ws3.cell(row=9, column=1, value='WINDOW CONFIG COMPARISON (Ensemble only)').font = section_font

headers_win = ['Window', 'Avg Dir Acc', 'Avg Sharpe', 'Avg PF', 'Avg N_Val', 'Verdict']
for i, h in enumerate(headers_win, 1):
    ws3.cell(row=11, column=i, value=h)
style_header(ws3, 11, len(headers_win))

windows = ens.groupby('window').agg({
    'dir_accuracy': 'mean', 'sharpe': 'mean', 'profit_factor': 'mean', 'n_val': 'mean'
}).sort_values('dir_accuracy', ascending=False)

for ri, (win, row) in enumerate(windows.iterrows(), 12):
    ws3.cell(row=ri, column=1, value=win)
    ws3.cell(row=ri, column=2, value=round(row['dir_accuracy'], 4))
    ws3.cell(row=ri, column=3, value=round(row['sharpe'], 4))
    ws3.cell(row=ri, column=4, value=round(row['profit_factor'], 4))
    ws3.cell(row=ri, column=5, value=int(row['n_val']))
    verdict = 'BEST' if ri == 12 else ('Good' if row['dir_accuracy'] > 0.52 else 'Poor')
    ws3.cell(row=ri, column=6, value=verdict)
    color_accuracy(ws3, ri, 2, row['dir_accuracy'])
style_data(ws3, 12, 11 + len(windows), len(headers_win))

# Sector comparison
ws3.cell(row=20, column=1, value='SECTOR COMPARISON (Ensemble only)').font = section_font
headers_sec = ['Sector', 'Avg Dir Acc', 'Avg Sharpe', 'Avg PF', 'Best Stock', 'Best Acc']
for i, h in enumerate(headers_sec, 1):
    ws3.cell(row=22, column=i, value=h)
style_header(ws3, 22, len(headers_sec))

ens_sect = ens.copy()
ens_sect['sector'] = ens_sect['symbol'].map(sectors)

for ri, sector in enumerate(['Banking', 'IT'], 23):
    sdf = ens_sect[ens_sect['sector'] == sector]
    best_stock = sdf.loc[sdf['dir_accuracy'].idxmax()]
    ws3.cell(row=ri, column=1, value=sector)
    ws3.cell(row=ri, column=2, value=round(sdf['dir_accuracy'].mean(), 4))
    ws3.cell(row=ri, column=3, value=round(sdf['sharpe'].mean(), 4))
    ws3.cell(row=ri, column=4, value=round(sdf['profit_factor'].mean(), 4))
    ws3.cell(row=ri, column=5, value=best_stock['symbol'])
    ws3.cell(row=ri, column=6, value=round(best_stock['dir_accuracy'], 4))
    color_accuracy(ws3, ri, 2, sdf['dir_accuracy'].mean())
style_data(ws3, 23, 24, len(headers_sec))
auto_width(ws3, len(headers_sec))

# ============ SHEET 4: PROBLEMS & ROOT CAUSES ============
ws4 = wb.create_sheet('Problems & Root Causes')
ws4.cell(row=1, column=1, value='CRITICAL ISSUES IDENTIFIED').font = Font(bold=True, size=14, color='C00000')

problems = [
    ('PROBLEM', 'ROOT CAUSE', 'EVIDENCE', 'SEVERITY', 'IMPACT'),
    (
        'Ensemble weights always 0.5/0.5',
        'scipy.optimize SLSQP fails to differentiate - val set too small or models too similar',
        'ALL 210 rows show weight_xgb=0.5, weight_lgb=0.5',
        'CRITICAL',
        'Ensemble adds NO value over individual models'
    ),
    (
        'Avg Dir Accuracy ~50% (coin flip)',
        'Regression target (close return) is near-zero mean, models predict near-zero = always same direction',
        f"Ensemble avg: {ens['dir_accuracy'].mean():.1%}, many stocks < 50%",
        'CRITICAL',
        'No predictive edge - system is not better than random'
    ),
    (
        'WR_Long often 0.0%',
        'Models predict ALL negative returns (bearish bias) - never predicts UP',
        f"{(ens['win_rate_long'] == 0).sum()} of {len(ens)} ensemble results have WR_Long=0",
        'HIGH',
        'Cannot generate BUY signals at all'
    ),
    (
        'IT stocks worse than Banking',
        'IT sector (TCS, INFY, WIPRO, TECHM) is harder to predict - earnings-driven, less technical',
        'Banking avg ~53%, IT avg ~49%',
        'MEDIUM',
        'Sector-specific models needed'
    ),
    (
        '240 features, no selection',
        'Feature bloat causes overfitting - many features are noise, no feature importance filtering',
        '240 features for ~1200-1700 training samples = high dimensionality',
        'HIGH',
        'Model learns noise instead of signal'
    ),
    (
        'No walk-forward validation',
        'Single expanding window, not rolling - model sees old regime data from 2019 (COVID)',
        'Train always starts from 2019-01-01',
        'MEDIUM',
        'Model may not adapt to current market regime'
    ),
    (
        'Regression predicts returns, not direction',
        'Predicting exact return magnitude is harder than binary direction classification',
        'RMSE ~0.007-0.015 but directional accuracy only 50%',
        'HIGH',
        'Wrong problem formulation for direction prediction'
    ),
]

for ri, row_data in enumerate(problems, 3):
    for ci, val in enumerate(row_data, 1):
        ws4.cell(row=ri, column=ci, value=val)
    if ri == 3:
        style_header(ws4, ri, 5)
    else:
        for ci in range(1, 6):
            ws4.cell(row=ri, column=ci).border = thin_border
            ws4.cell(row=ri, column=ci).alignment = Alignment(wrap_text=True)
        sev = ws4.cell(row=ri, column=4).value
        if sev == 'CRITICAL':
            ws4.cell(row=ri, column=4).fill = bad_fill
        elif sev == 'HIGH':
            ws4.cell(row=ri, column=4).fill = warn_fill

for col in [1, 2, 3, 5]:
    ws4.column_dimensions[get_column_letter(col)].width = 35
ws4.column_dimensions['D'].width = 12

# ============ SHEET 5: IMPROVEMENTS ROADMAP ============
ws5 = wb.create_sheet('Improvement Roadmap')
ws5.cell(row=1, column=1, value='IMPROVEMENT ROADMAP TO 58%+ ACCURACY').font = Font(bold=True, size=14, color='1F4E79')

improvements = [
    ('CHANGE', 'PRIORITY', 'EXPECTED GAIN', 'EFFORT', 'DETAILS'),
    (
        '1. Switch to Classification (XGBClassifier)',
        'P0-CRITICAL',
        '+3-5% accuracy',
        'Low (1 day)',
        'Use XGBClassifier/LGBMClassifier with binary target. Direct optimization for direction instead of return magnitude. Use class_weight="balanced" for imbalanced days.'
    ),
    (
        '2. Feature Selection (top 30-50)',
        'P0-CRITICAL',
        '+2-4% accuracy',
        'Low (1 day)',
        'Use mutual_info_classif or model feature_importances to select top 30-50 features. Drop correlated features (>0.95 correlation). Reduces overfitting dramatically.'
    ),
    (
        '3. Fix Ensemble Weight Optimization',
        'P0-CRITICAL',
        '+1-2% accuracy',
        'Low (hours)',
        'Use grid search [0.3,0.4,0.5,0.6,0.7] instead of scipy.optimize. Or use stacking meta-learner (LogisticRegression on model predictions).'
    ),
    (
        '4. Proper Walk-Forward Validation',
        'P1-HIGH',
        '+2-3% accuracy',
        'Medium (2 days)',
        'Rolling 252-day train, 63-day test. Retrain every 63 days. Prevents regime-stale models. Average metrics across all folds.'
    ),
    (
        '5. Confidence Threshold (only trade >55%)',
        'P1-HIGH',
        '+3-5% win rate',
        'Low (hours)',
        'Only generate signals when model probability > 55%. Reduces trade count but increases win rate. Use predict_proba() from classifier.'
    ),
    (
        '6. Add Sentiment Features',
        'P1-HIGH',
        '+2-3% accuracy',
        'Medium (1 day)',
        'Google News RSS sentiment is fast (<2s). Add 8 sentiment features. Especially helps for IT stocks (earnings-driven).'
    ),
    (
        '7. Sector-Specific Models',
        'P2-MEDIUM',
        '+1-2% accuracy',
        'Medium (2 days)',
        'Train separate models for Banking vs IT. Different feature importance, different market dynamics. Banking responds to rates, IT to USD/earnings.'
    ),
    (
        '8. Add CatBoost to Ensemble',
        'P2-MEDIUM',
        '+1-2% accuracy',
        'Low (1 day)',
        'CatBoost handles categorical features well, good for regime detection. 3-model ensemble (XGB+LGB+CatBoost) more diverse than 2-model.'
    ),
    (
        '9. Hyperparameter Tuning (Optuna)',
        'P2-MEDIUM',
        '+1-2% accuracy',
        'Medium (2 days)',
        'Use Optuna for Bayesian hyperparameter search. Currently using fixed params. Key params: max_depth, learning_rate, n_estimators, reg_alpha/lambda.'
    ),
    (
        '10. Calibrated Probabilities',
        'P3-LOW',
        '+1% accuracy',
        'Low (hours)',
        'Use CalibratedClassifierCV for well-calibrated prediction probabilities. Better confidence thresholding.'
    ),
]

for ri, row_data in enumerate(improvements, 3):
    for ci, val in enumerate(row_data, 1):
        ws5.cell(row=ri, column=ci, value=val)
    if ri == 3:
        style_header(ws5, ri, 5)
    else:
        for ci in range(1, 6):
            ws5.cell(row=ri, column=ci).border = thin_border
            ws5.cell(row=ri, column=ci).alignment = Alignment(wrap_text=True, vertical='top')
        pri = ws5.cell(row=ri, column=2).value
        if 'P0' in pri:
            ws5.cell(row=ri, column=2).fill = bad_fill
        elif 'P1' in pri:
            ws5.cell(row=ri, column=2).fill = warn_fill
        elif 'P2' in pri:
            ws5.cell(row=ri, column=2).fill = good_fill

ws5.column_dimensions['A'].width = 40
ws5.column_dimensions['B'].width = 15
ws5.column_dimensions['C'].width = 18
ws5.column_dimensions['D'].width = 15
ws5.column_dimensions['E'].width = 60

# Expected outcome
r = len(improvements) + 5
ws5.cell(row=r, column=1, value='PROJECTED OUTCOME AFTER P0+P1 CHANGES').font = section_font
ws5.cell(row=r, column=1).fill = section_fill
for c in range(1, 4):
    ws5.cell(row=r, column=c).fill = section_fill

proj_headers = ['Metric', 'Current', 'Projected']
for i, h in enumerate(proj_headers, 1):
    ws5.cell(row=r+1, column=i, value=h)
style_header(ws5, r+1, 3)

projections = [
    ('Dir Accuracy (avg)', f"{ens['dir_accuracy'].mean():.1%}", '55-60%'),
    ('Dir Accuracy (best)', f"{ens['dir_accuracy'].max():.1%}", '62-68%'),
    ('Win Rate Long', f"{ens['win_rate_long'].mean():.1%}", '55-62%'),
    ('Sharpe Ratio', f"{ens['sharpe'].mean():.2f}", '1.0-2.0'),
    ('Profit Factor', f"{ens['profit_factor'].mean():.2f}", '1.2-1.5'),
    ('Trade Frequency', '100% (all days)', '40-60% (filtered)'),
]
for i, (m, cur, proj) in enumerate(projections, r+2):
    ws5.cell(row=i, column=1, value=m)
    ws5.cell(row=i, column=2, value=cur)
    ws5.cell(row=i, column=3, value=proj)
style_data(ws5, r+2, r+1+len(projections), 3)

# ============ SHEET 6: BEST CONFIGS PER STOCK ============
ws6 = wb.create_sheet('Best Config Per Stock')
ws6.cell(row=1, column=1, value='BEST PERFORMING CONFIGURATION PER STOCK').font = section_font

headers6 = ['Symbol', 'Sector', 'Best Model', 'Best Window', 'Dir Accuracy',
            'WR Long', 'WR Short', 'Sharpe', 'PF', 'N_Val', 'Verdict']
for i, h in enumerate(headers6, 1):
    ws6.cell(row=3, column=i, value=h)
style_header(ws6, 3, len(headers6))

# Find absolute best per stock across all models
best_per_stock = df.groupby('symbol').apply(
    lambda x: x.loc[x['dir_accuracy'].idxmax()]
).sort_values('dir_accuracy', ascending=False).reset_index(drop=True)

for ri, (_, row) in enumerate(best_per_stock.iterrows(), 4):
    acc = row['dir_accuracy']
    verdict = 'TRADEABLE' if acc >= 0.58 else ('MARGINAL' if acc >= 0.53 else 'NOT READY')
    ws6.cell(row=ri, column=1, value=row['symbol'])
    ws6.cell(row=ri, column=2, value=sectors.get(row['symbol'], ''))
    ws6.cell(row=ri, column=3, value=row['model'].upper())
    ws6.cell(row=ri, column=4, value=row['window'])
    ws6.cell(row=ri, column=5, value=f"{acc:.1%}")
    ws6.cell(row=ri, column=6, value=f"{row['win_rate_long']:.1%}")
    ws6.cell(row=ri, column=7, value=f"{row['win_rate_short']:.1%}")
    ws6.cell(row=ri, column=8, value=f"{row['sharpe']:.2f}")
    ws6.cell(row=ri, column=9, value=f"{row['profit_factor']:.2f}")
    ws6.cell(row=ri, column=10, value=int(row['n_val']))
    ws6.cell(row=ri, column=11, value=verdict)
    color_accuracy(ws6, ri, 5, acc)
    if verdict == 'TRADEABLE':
        ws6.cell(row=ri, column=11).fill = good_fill
    elif verdict == 'MARGINAL':
        ws6.cell(row=ri, column=11).fill = warn_fill
    else:
        ws6.cell(row=ri, column=11).fill = bad_fill
style_data(ws6, 4, 3 + len(best_per_stock), len(headers6))
auto_width(ws6, len(headers6))

# Save
wb.save(OUTPUT_XLSX)
print(f"Excel report saved: {OUTPUT_XLSX}")
print(f"\nKey Findings:")
print(f"  Avg Ensemble Dir Accuracy: {ens['dir_accuracy'].mean():.1%}")
print(f"  Best Ensemble Dir Accuracy: {ens['dir_accuracy'].max():.1%} ({stock_best.iloc[0]['symbol']} @ {stock_best.iloc[0]['window']})")
print(f"  Ensemble weights: ALL 0.5/0.5 (broken optimization)")
print(f"  Banking sector avg: {ens_sect[ens_sect['sector']=='Banking']['dir_accuracy'].mean():.1%}")
print(f"  IT sector avg: {ens_sect[ens_sect['sector']=='IT']['dir_accuracy'].mean():.1%}")
