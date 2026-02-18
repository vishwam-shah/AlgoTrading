"""
================================================================================
CREATE EXCEL REPORTS - Professional Excel Output with Formatting
================================================================================
Export model comparison results to properly formatted Excel files with:
- Multiple sheets (summary, detailed metrics, charts)
- Professional formatting (colors, borders, fonts)
- Charts and visualizations
- Conditional formatting for performance metrics
================================================================================
"""

import pandas as pd
import numpy as np
from pathlib import Path
import sys

# Add paths
v3_path = Path(__file__).parent.parent
sys.path.insert(0, str(v3_path))

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference, LineChart
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("[WARNING] openpyxl not installed. Install with: pip install openpyxl")


def create_formatted_excel_report(csv_path, output_path):
    """
    Create a professionally formatted Excel report from CSV results.

    Args:
        csv_path: Path to CSV file with model comparison results
        output_path: Path to save Excel file
    """
    print("="*70)
    print(" CREATING EXCEL REPORT")
    print("="*70)

    # Load data
    print(f"\n[1/5] Loading data from: {csv_path}")
    df = pd.read_csv(csv_path)

    if not OPENPYXL_AVAILABLE:
        # Fallback: Simple Excel without formatting
        print("\n[WARNING] Creating simple Excel without formatting (openpyxl not available)")
        df.to_excel(output_path, index=False, sheet_name='Model Comparison')
        print(f"\n[COMPLETE] Basic Excel saved to: {output_path}")
        return

    # Create Excel writer
    print(f"\n[2/5] Creating Excel workbook...")
    writer = pd.ExcelWriter(output_path, engine='openpyxl')

    # Sheet 1: Summary Rankings
    print(f"\n[3/5] Creating Summary sheet...")
    summary_df = df[['Model', 'Type', 'Accuracy', 'F1', 'TrainTime']].copy()
    summary_df['Accuracy'] = summary_df['Accuracy'] * 100  # Convert to percentage
    summary_df['Rank'] = range(1, len(summary_df) + 1)
    summary_df = summary_df[['Rank', 'Model', 'Type', 'Accuracy', 'F1', 'TrainTime']]
    summary_df.to_excel(writer, sheet_name='Summary', index=False)

    # Sheet 2: Detailed Metrics
    print(f"\n[4/5] Creating Detailed Metrics sheet...")
    detailed_df = df.copy()
    detailed_df['Accuracy'] = detailed_df['Accuracy'] * 100
    detailed_df['Precision'] = detailed_df['Precision'] * 100
    detailed_df['Recall'] = detailed_df['Recall'] * 100
    detailed_df.to_excel(writer, sheet_name='Detailed Metrics', index=False)

    # Sheet 3: Model Type Comparison
    print(f"\n[5/5] Creating Type Comparison sheet...")
    type_summary = df.groupby('Type').agg({
        'Accuracy': ['mean', 'max', 'min'],
        'F1': 'mean',
        'TrainTime': 'mean'
    }).round(4)
    type_summary.columns = ['Avg_Accuracy', 'Max_Accuracy', 'Min_Accuracy', 'Avg_F1', 'Avg_TrainTime']
    type_summary = type_summary.reset_index()
    type_summary['Avg_Accuracy'] = type_summary['Avg_Accuracy'] * 100
    type_summary['Max_Accuracy'] = type_summary['Max_Accuracy'] * 100
    type_summary['Min_Accuracy'] = type_summary['Min_Accuracy'] * 100
    type_summary.to_excel(writer, sheet_name='Type Comparison', index=False)

    writer.close()

    # Apply formatting
    print(f"\n[FORMATTING] Applying professional formatting...")
    wb = load_workbook(output_path)

    # Format Summary sheet
    ws_summary = wb['Summary']
    format_summary_sheet(ws_summary)

    # Format Detailed Metrics sheet
    ws_detailed = wb['Detailed Metrics']
    format_detailed_sheet(ws_detailed)

    # Format Type Comparison sheet
    ws_type = wb['Type Comparison']
    format_type_comparison_sheet(ws_type)

    # Add charts
    add_comparison_charts(wb, df)

    # Save
    wb.save(output_path)

    print(f"\n{'='*70}")
    print(f" [SUCCESS] Excel report created: {output_path}")
    print(f"{'='*70}")
    print(f"\nSheets created:")
    print(f"  1. Summary - Model rankings")
    print(f"  2. Detailed Metrics - All performance metrics")
    print(f"  3. Type Comparison - Traditional ML vs Deep Learning")
    print(f"  4. Charts - Visual comparisons")


def format_summary_sheet(ws):
    """Format the summary sheet with colors and styling."""
    # Header styling
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Data styling
    for row in range(2, ws.max_row + 1):
        # Rank column - bold
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].alignment = Alignment(horizontal='center')

        # Model column - bold
        ws[f'B{row}'].font = Font(bold=True)

        # Accuracy column - conditional formatting
        accuracy = ws[f'D{row}'].value
        if accuracy >= 55:
            ws[f'D{row}'].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            ws[f'D{row}'].font = Font(color='006100')
        elif accuracy >= 52:
            ws[f'D{row}'].fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            ws[f'D{row}'].font = Font(color='9C5700')
        else:
            ws[f'D{row}'].fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            ws[f'D{row}'].font = Font(color='9C0006')

        # Number formatting
        ws[f'D{row}'].number_format = '0.00"%"'
        ws[f'E{row}'].number_format = '0.0000'
        ws[f'F{row}'].number_format = '0.0"s"'

    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12

    # Borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.border = thin_border


def format_detailed_sheet(ws):
    """Format the detailed metrics sheet."""
    # Header styling
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Number formatting for percentage columns
    percentage_cols = ['C', 'D', 'E']  # Accuracy, Precision, Recall
    for col in percentage_cols:
        for row in range(2, ws.max_row + 1):
            ws[f'{col}{row}'].number_format = '0.00"%"'

    # F1 and AUC formatting
    for row in range(2, ws.max_row + 1):
        ws[f'F{row}'].number_format = '0.0000'
        ws[f'G{row}'].number_format = '0.0000'
        ws[f'H{row}'].number_format = '0.0000'
        ws[f'I{row}'].number_format = '0.0"s"'

    # Column widths
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col].width = 14


def format_type_comparison_sheet(ws):
    """Format the type comparison sheet."""
    # Header styling
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Type column - bold
    for row in range(2, ws.max_row + 1):
        ws[f'A{row}'].font = Font(bold=True)

        # Format percentages
        for col in ['B', 'C', 'D']:
            ws[f'{col}{row}'].number_format = '0.00"%"'

        ws[f'E{row}'].number_format = '0.0000'
        ws[f'F{row}'].number_format = '0.0"s"'

    # Column widths
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 16


def add_comparison_charts(wb, df):
    """Add charts to visualize model comparison."""
    # Create new sheet for charts
    ws_charts = wb.create_sheet('Charts')

    # Chart 1: Accuracy comparison (Bar chart)
    chart1 = BarChart()
    chart1.title = "Model Accuracy Comparison"
    chart1.x_axis.title = "Model"
    chart1.y_axis.title = "Accuracy (%)"
    chart1.width = 15
    chart1.height = 10

    ws_summary = wb['Summary']

    # Data for chart (Models and Accuracy)
    data = Reference(ws_summary, min_col=4, min_row=1, max_row=ws_summary.max_row)
    categories = Reference(ws_summary, min_col=2, min_row=2, max_row=ws_summary.max_row)

    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(categories)

    ws_charts.add_chart(chart1, "A1")

    # Chart 2: Training Time comparison
    chart2 = BarChart()
    chart2.title = "Training Time Comparison"
    chart2.x_axis.title = "Model"
    chart2.y_axis.title = "Time (seconds)"
    chart2.width = 15
    chart2.height = 10

    time_data = Reference(ws_summary, min_col=6, min_row=1, max_row=ws_summary.max_row)
    chart2.add_data(time_data, titles_from_data=True)
    chart2.set_categories(categories)

    ws_charts.add_chart(chart2, "A25")


def main():
    """Main function to create Excel reports."""
    print("\n" + "="*70)
    print(" EXCEL REPORT GENERATOR")
    print("="*70)

    # Paths
    csv_path = v3_path / '06_results' / 'all_models_comparison.csv'
    output_path = v3_path / '06_results' / 'MODEL_COMPARISON_REPORT.xlsx'

    if not csv_path.exists():
        print(f"\n[ERROR] CSV file not found: {csv_path}")
        print("Please run test_all_models.py first to generate results.")
        return

    # Create report
    create_formatted_excel_report(csv_path, output_path)

    print(f"\n[COMPLETE] Professional Excel report ready!")
    print(f"\nOpen file: {output_path}")


if __name__ == '__main__':
    main()
